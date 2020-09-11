from openpyxl import load_workbook
import requests
import os

caminho = os.listdir("./Pastas_para_Importacao")
pasta = input("Insira a pasta que está a tabela:")
arquivo = input("Insira o nome do arquivo:")

wb = load_workbook(f'./%s/%s.xlsx' %(pasta, arquivo))
ws = wb.active

max_linha = ws.max_row
max_col = ws.max_column

for i in range (2, max_linha+1):
    cellUrl = ws.cell(row=i, column = 2)
    cellCod = ws.cell(row=i, column = 1)

    url = cellUrl.value
    cod = cellCod.value
    codi = (f"(%s)" %cod)
    
    for cam in caminho:
        cami = cam[-9::]
        if(cami == codi):
            print(cam)
    
            try:
        
                r = requests.get(url)
                with open(f'./Pastas_para_Importacao/%s/(%s).jpg' %(cam, cam), 'wb') as f:
                    f.write(r.content)
            except requests.exceptions.HTTPError as e:
                print("Error: " + str(e))
